from openpyxl.utils import get_column_letter,column_index_from_string
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\sungue\Desktop\试剂柜张贴表-最终版.xlsx')
sheets = wb.sheetnames
print(sheets)
print(wb.active.title)

sheet = wb.active
print(sheet['a1'].value)
cell = sheet['b5']
print(cell.row,cell.column,cell.coordinate)

cell = sheet.cell(row = 4,column = 6)
print(cell.value,cell.coordinate)

print(sheet.max_row,sheet.max_column)       #获取最大行和最大列的数值


print(get_column_letter(45),column_index_from_string('gfr'))          #列的数字与字母的转换


agent = sheet['a1':'f8']
print (list(agent))
for i in agent:
	for j in i:
		print(j.value,end = ',')
	print()
	print('***********************')

print(list(sheet.columns))
for i in list(sheet.columns)[0]:
	print(i.value)
