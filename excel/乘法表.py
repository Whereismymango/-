def caluate():
	n = int(input('请输入一个数字： '))
	import openpyxl,os
	from openpyxl.styles import Font,colors,Alignment
	os.chdir(r'C:\Intel\SGC\新建文件夹\excel')
	wb = openpyxl.Workbook()
	sheet = wb.active
	font = Font(name = '宋体',bold = True,italic = True,size = 15,color = colors.BLUE)
	for i in range(1,n+1):
		sheet.cell(row = 1,column = i+1,value = i).font = font
		
	for i in range(1,n+1):
		sheet.cell(column = 1,row = i+1,value = i).font = font
	
	
	
	
	for i in range(2,n+2):
		for j in range(2,n+2):
			sheet.cell(row = i,column = j,value = (i-1)*(j-1))
			
	wb.save('乘法表.xlsx')

caluate()