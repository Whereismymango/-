import openpyxl,os
os.chdir(r'C:\Intel\SGC\新建文件夹\excel')
wb = openpyxl.load_workbook('试剂柜张贴表-最终版.xlsx')
wb1 = openpyxl.Workbook()
names = wb.sheetnames
x = 0

for name in names[1:]:
	y = 0
	x += 1	
	a = wb1.create_sheet(name,index = x)
	sheet = wb[name]
	for row in sheet.values:
		y += 1
		a.append(list(row))	
		if list(row)[-1] == None:
			a.merge_cells('A'+str(y)+':'+'F'+str(y))
		else:
			pass
wb1.save('wb.xlsx')
