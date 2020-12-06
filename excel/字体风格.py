from openpyxl.styles import Font,colors,Alignment
import openpyxl,os
os.chdir(r'C:\Intel\SGC\新建文件夹\excel')
wb = openpyxl.Workbook()
sheet = wb.active
font = Font(name='宋体',bold=True,italic=True,color=colors.RED)   #颜色必须大写字母
sheet['a1'].font = font
sheet['a1'].value = '我正在学python'             #这里有没有.value都可以
sheet['a1'].alignment = Alignment(horizontal = 'right',vertical = 'top')
sheet.row_dimensions[1].height = 0
sheet.column_dimensions['a'].width = 20

sheet.merge_cells('a2:e5')
sheet.unmerge_cells('a2:e5')



wb.save('q.xlsx')
