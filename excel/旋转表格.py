import openpyxl,os
from openpyxl.utils import column_index_from_string,get_column_letter
os.chdir(r'C:\Intel\SGC\新建文件夹\excel')
a = []
wb = openpyxl.load_workbook('q.xlsx')
wb1 = openpyxl.Workbook()
sheet = wb.active
sheet1 = wb1.active
for i in range(1,sheet.max_row+1):
	b = []
	for j in range(1,sheet.max_column+1):
		b.append(sheet.cell(row = i,column = j).value)
	a.append(b)	
	
	
for i in range(1,sheet.max_column+1):
	for j in range(1,sheet.max_row+1):
		sheet1.cell(row = i,column = j,value = a[j-1][i-1])
		
wb1.save('旋转.xlsx')