from openpyxl.styles import Font,colors,Alignment
import openpyxl,os
os.chdir(r'C:\Intel\SGC\新建文件夹\excel')
wb = openpyxl.Workbook()
sheet = wb.active
#font = Font(name='宋体',size=24,bold=True,italic=True,color=colors.RED)   #颜色必须大写字母
#sheet['a1'].font = font
#sheet['a1'].value = 'dhsubvds899成都市不会v看'

for i in  range(1,5):
	for j in range(1,5):
		sheet.cell(row=i, column=j,value= 2*i+3*j)     #访问的同时写入value值

sheet.append([1, 2, 3])
sheet.append(list(range(100)))
sheet1 = wb.create_sheet('huih',0)     #初始化的同时也同时新建新工作表

s = sheet['C:D']
#print(s)

sheet.sheet_properties.tabColor = '4C2E28'
sheet1.sheet_properties.tabColor = '40ff00'
#wb.remove(sheet)
#del wb[wb.sheetnames[0]]

rows = [
    ['Number', 'data1', 'data2'],
    [40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10]]
	
a = list(zip(*rows))
print(a)
wb.save('q.xlsx')
