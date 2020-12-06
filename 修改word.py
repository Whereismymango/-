from openpyxl import *
wb = load_workbook(r'C:\Users\sungue\Desktop\危险化学品清单-07.10.2020.xlsx')
sheet = wb.active
print(sheet.title)
a = {}
for i in range(1,sheet.max_row+1):
	if sheet.cell(row = i,column = 1).value == None:
		pass
	elif sheet.cell(row = i,column = 2).value == None:
		pass
	else:
		a[sheet.cell(row = i,column = 1).value] = sheet.cell(row = i,column = 2).value
#print(a)
#print(len(a))	


from docx import Document
import os
os.chdir(r'C:\Intel\SGC\试剂')
document = Document('原辅料组实验准备信息指南 Third_edition.docx')
tables = document.tables
#print(tables)
print(len(tables))

q = 1
for table in tables[2::2]:
	c = len(table.columns)
	print('这是第'+str(q)+'个表')
	for i in range(len(table.rows)):
		result = table.cell(i,c-2).text
		if result in a:
			table.cell(i,c-2).text = a[result]
		print(table.cell(i,c-2).text)
	print('\n\n')
	q += 1
document.save('1.docx')
