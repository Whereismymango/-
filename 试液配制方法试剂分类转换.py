from openpyxl import *
wb = load_workbook(r'C:\Users\sungue\Desktop\危险化学品清单-07.10.2020.xlsx')
sheet = wb.active
print(sheet.title)
a = {}
for i in range(1,sheet.max_row+1):
	if sheet.cell(row = i,column = 1).value == None:
		pass
	else:
		a[sheet.cell(row = i,column = 1).value] = sheet.cell(row = i,column = 2).value
print(a)
#print(len(a))	




wb1 = load_workbook(r'C:\Users\sungue\Desktop\试液配置方法.xlsx')
sheet1 = wb1.active
for i in sheet1['e']:
	if i.value in a:
		i.value = a[i.value]
	else:
		pass
		
		
wb1.save(r'C:\Users\sungue\Desktop\试液配置方法_copy.xlsx')